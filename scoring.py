from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd

RAW_INPUT_HEADERS = [
    "S.No.",
    "Call Date",
    "Audit Date",
    "Quality Auditor",
    "Call/Chat",
    "Agent Name",
    "Supervisor name",
    "LOB",
    "Sub-LOB",
    "Mobile Number",
    "Reason for call",
    "Agent's Response",
    "Greeting",
    "Permission",
    "Closing",
    "Active Listening",
    "Empathy/Apology",
    "Politeness",
    "Preferred Language",
    "Voice Clarity/Tone",
    "Context Setting",
    "Grammar & Sentence",
    "Probing",
    "Correct Resolution",
    "Complete Resolution",
    "TAT informed",
    "Hold (<=1min 3x)",
    "Transfer/Escalation",
    "Complain/Ticket Raised",
    "Response Time (<10s)",
    "All Queries Tagged",
    "Complete Tagging",
    "Correct Tagging",
    "Upsell/Promotions",
    "Waiver/Discount",
    "Condescending/Rude/Abuse",
    "Disconnect Line",
    "Personal Info Violation",
    "Complaints tagged wrongly",
    "Blind Transfer",
    "Escalation Denied",
]

RAW_SECTION_COLUMNS = {
    "Call Compliance": ["Greeting", "Permission", "Closing", "Hold (<=1min 3x)", "Transfer/Escalation", "Complain/Ticket Raised", "Response Time (<10s)"],
    "Call Etiquette": [
        "Active Listening",
        "Empathy/Apology",
        "Politeness",
        "Preferred Language",
        "Voice Clarity/Tone",
        "Context Setting",
        "Grammar & Sentence",
    ],
    "Query Resolution": ["Probing", "Correct Resolution", "Complete Resolution", "TAT informed"],
    "Disposition": ["All Queries Tagged", "Complete Tagging", "Correct Tagging"],
    "Upsell/Promotions": ["Upsell/Promotions"],
    "Empowerment": ["Waiver/Discount"],
}

COMPLIANCE_FLAG_COLUMNS = [
    "Condescending/Rude/Abuse",
    "Disconnect Line",
    "Personal Info Violation",
    "Complaints tagged wrongly",
    "Blind Transfer",
    "Escalation Denied",
]


@dataclass(frozen=True)
class Parameter:
    section: str
    name: str
    raw_col: str
    max_score: int
    kind: str
    fatal_enabled: bool = False


PARAMETERS: list[Parameter] = [
    Parameter("Call Compliance", "Greeting", "Greeting", 2, "binary", True),
    Parameter("Call Compliance", "Permission", "Permission", 2, "binary"),
    Parameter("Call Compliance", "Closing", "Closing", 2, "binary"),
    Parameter("Call Etiquette", "Active Listening", "Active Listening", 6, "qualitative"),
    Parameter("Call Etiquette", "Empathy/Apology", "Empathy/Apology", 6, "qualitative"),
    Parameter("Call Etiquette", "Politeness", "Politeness", 6, "qualitative"),
    Parameter("Call Etiquette", "Preferred Language", "Preferred Language", 3, "qualitative"),
    Parameter("Call Etiquette", "Voice Clarity/Tone", "Voice Clarity/Tone", 3, "qualitative"),
    Parameter("Call Etiquette", "Context Setting", "Context Setting", 3, "qualitative"),
    Parameter("Call Etiquette", "Grammar & Sentence", "Grammar & Sentence", 6, "qualitative"),
    Parameter("Query Resolution", "Probing", "Probing", 6, "qualitative"),
    Parameter("Query Resolution", "Correct Resolution", "Correct Resolution", 10, "binary", True),
    Parameter("Query Resolution", "Complete Resolution", "Complete Resolution", 10, "binary", True),
    Parameter("Query Resolution", "TAT informed", "TAT informed", 3, "binary"),
    Parameter("Call Compliance", "Hold (<=1min 3x)", "Hold (<=1min 3x)", 3, "binary"),
    Parameter("Call Compliance", "Transfer/Escalation", "Transfer/Escalation", 2, "binary"),
    Parameter("Call Compliance", "Complain/Ticket Raised", "Complain/Ticket Raised", 3, "binary", True),
    Parameter("Call Compliance", "Response Time (<10s)", "Response Time (<10s)", 4, "binary"),
    Parameter("Disposition", "All Queries Tagged", "All Queries Tagged", 4, "binary"),
    Parameter("Disposition", "Complete Tagging", "Complete Tagging", 4, "binary", True),
    Parameter("Disposition", "Correct Tagging", "Correct Tagging", 6, "binary", True),
    Parameter("Upsell/Promotions", "Upsell/Promotions", "Upsell/Promotions", 4, "binary"),
    Parameter("Empowerment", "Waiver/Discount", "Waiver/Discount", 2, "binary"),
]

DISPLAY_COLUMNS = [
    "Record ID",
    "Call Date",
    "Audit Date",
    "Quality Auditor",
    "Call/Chat",
    "Agent Name",
    "Supervisor name",
    "LOB",
    "Sub-LOB",
    "Reason for call",
    "Overall Score (With Fatal)",
    "Overall Score (Without Fatal)",
    "Defect %",
    "Fatal Count",
    "Issue Summary",
]


def normalize_response(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and np.isnan(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    return text.upper()


def parse_score(response: Any, max_score: int, kind: str, fatal_enabled: bool = False) -> dict[str, Any]:
    normalized = normalize_response(response)
    result = {
        "normalized": normalized,
        "applicable": False,
        "score": np.nan,
        "pct": np.nan,
        "fatal": False,
        "failed": False,
        "partial": False,
        "status": "Excluded",
    }

    if normalized in {None, "NA"}:
        return result

    result["applicable"] = True

    if fatal_enabled and normalized == "FATAL":
        result.update({"score": 0.0, "pct": 0.0, "fatal": True, "failed": True, "status": "Fatal"})
        return result

    if kind == "binary":
        if normalized == "Y":
            score = float(max_score)
        elif normalized == "N":
            score = 0.0
        else:
            result.update({"applicable": False})
            return result
    elif kind == "qualitative":
        if normalized == "EE":
            score = float(max_score)
        elif normalized == "ME":
            score = 3.0 if max_score == 6 else 2.0
        elif normalized == "BE":
            score = 0.0
        else:
            result.update({"applicable": False})
            return result
    else:
        raise ValueError(f"Unsupported parameter kind: {kind}")

    pct = score / max_score if max_score else np.nan
    failed = score == 0
    partial = 0 < score < max_score
    status = "Pass"
    if failed:
        status = "Fail"
    elif partial:
        status = "Partial"

    result.update({"score": score, "pct": pct, "failed": failed, "partial": partial, "status": status})
    return result


def _finalize_input_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.dropna(how="all")
    key_cols = ["S.No.", "Agent Name", "Call Date", "Audit Date"]
    df = df.loc[df[key_cols].notna().any(axis=1)].copy()
    for col in ["Call Date", "Audit Date"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    return df.reset_index(drop=True)


def raw_rows_to_dataframe(rows: list[list[Any]] | list[tuple[Any, ...]], start_row: int = 3) -> pd.DataFrame:
    normalized_rows: list[dict[str, Any]] = []
    max_len = len(RAW_INPUT_HEADERS)
    for idx, values in enumerate(rows, start=start_row):
        values = list(values[:max_len]) + [None] * max(0, max_len - len(values))
        if not any(value not in (None, "") for value in values):
            continue
        row = {RAW_INPUT_HEADERS[col_idx]: values[col_idx] for col_idx in range(max_len)}
        row["_excel_row"] = idx
        normalized_rows.append(row)
    return _finalize_input_dataframe(pd.DataFrame(normalized_rows))


def _excel_to_dataframe(workbook_source: str | Path | BytesIO | bytes) -> pd.DataFrame:
    if isinstance(workbook_source, bytes):
        workbook_source = BytesIO(workbook_source)
    wb = openpyxl.load_workbook(workbook_source, data_only=False, read_only=True)
    if "Audit sheet" not in wb.sheetnames:
        raise ValueError("Workbook does not contain a sheet named 'Audit sheet'.")
    ws = wb["Audit sheet"]
    rows: list[dict[str, Any]] = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=3, max_col=41, values_only=True), start=3):
        if not any(value not in (None, "") for value in values):
            continue
        row = {RAW_INPUT_HEADERS[idx]: values[idx] for idx in range(len(RAW_INPUT_HEADERS))}
        row["_excel_row"] = excel_row
        rows.append(row)
    return _finalize_input_dataframe(pd.DataFrame(rows))


def _issue_summary(row: pd.Series) -> str:
    findings: list[str] = []
    for parameter in PARAMETERS:
        status = row[f"status__{parameter.name}"]
        raw = row.get(parameter.raw_col)
        if status in {"Fail", "Fatal", "Partial"}:
            findings.append(f"{parameter.name}: {raw}")
    return " | ".join(findings) if findings else "Clean audit"


def _build_scored_dataset_from_df(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if df.empty:
        return df, pd.DataFrame(), pd.DataFrame()

    df["Record ID"] = df["S.No."].apply(lambda x: f"AUD-{int(x):04d}" if pd.notna(x) else None)
    df["Month"] = df["Call Date"].dt.to_period("M").astype(str)
    df["Week"] = df["Call Date"].dt.to_period("W").astype(str)
    df["Audit Month"] = df["Audit Date"].dt.to_period("M").astype(str)

    long_rows: list[dict[str, Any]] = []
    compliance_rows: list[dict[str, Any]] = []
    derived_columns: dict[str, pd.Series] = {}

    for parameter in PARAMETERS:
        parsed = df[parameter.raw_col].apply(lambda value: parse_score(value, parameter.max_score, parameter.kind, parameter.fatal_enabled))
        derived_columns[f"score__{parameter.name}"] = parsed.apply(lambda item: item["score"])
        derived_columns[f"pct__{parameter.name}"] = parsed.apply(lambda item: item["pct"])
        derived_columns[f"status__{parameter.name}"] = parsed.apply(lambda item: item["status"])
        derived_columns[f"fatal__{parameter.name}"] = parsed.apply(lambda item: item["fatal"])
        derived_columns[f"failed__{parameter.name}"] = parsed.apply(lambda item: item["failed"])
        derived_columns[f"applicable__{parameter.name}"] = parsed.apply(lambda item: item["applicable"])

    df = pd.concat([df, pd.DataFrame(derived_columns)], axis=1)

    for section, parameter_names in RAW_SECTION_COLUMNS.items():
        section_parameters = [p for p in PARAMETERS if p.name in parameter_names]
        score_cols = [f"score__{p.name}" for p in section_parameters]
        applicable_cols = [f"applicable__{p.name}" for p in section_parameters]
        max_scores = [p.max_score for p in section_parameters]
        section_earned = df[score_cols].fillna(0).sum(axis=1)
        section_denom = sum(df[col].astype(int) * max_score for col, max_score in zip(applicable_cols, max_scores))
        df[f"Section Score :: {section}"] = np.where(section_denom > 0, section_earned / section_denom, np.nan)

    applicable_cols_all = [f"applicable__{p.name}" for p in PARAMETERS]
    score_cols_all = [f"score__{p.name}" for p in PARAMETERS]
    fatal_cols_all = [f"fatal__{p.name}" for p in PARAMETERS]
    failed_cols_all = [f"failed__{p.name}" for p in PARAMETERS]

    total_possible = sum(df[col].astype(int) * p.max_score for col, p in zip(applicable_cols_all, PARAMETERS))
    total_earned = df[score_cols_all].fillna(0).sum(axis=1)
    fatal_count = df[fatal_cols_all].sum(axis=1)
    failed_count = df[failed_cols_all].sum(axis=1)
    applicable_count = df[applicable_cols_all].sum(axis=1)

    df["Fatal Count"] = fatal_count
    df["Failed Parameter Count"] = failed_count
    df["Applicable Parameter Count"] = applicable_count
    df["Overall Score (Without Fatal)"] = np.where(total_possible > 0, total_earned / total_possible, np.nan)
    df["Overall Score (With Fatal)"] = np.where(fatal_count > 0, 0.0, df["Overall Score (Without Fatal)"])
    df["Defect %"] = np.where(applicable_count > 0, failed_count / applicable_count, np.nan)

    for flag_col in COMPLIANCE_FLAG_COLUMNS:
        normalized = df[flag_col].apply(normalize_response)
        df[f"Compliance Flag :: {flag_col}"] = normalized
        df[f"Compliance Flag Positive :: {flag_col}"] = normalized.isin({"Y", "YES", "TRUE", "1"})

    positive_flag_cols = [f"Compliance Flag Positive :: {flag_col}" for flag_col in COMPLIANCE_FLAG_COLUMNS]
    df["Compliance Positive Count"] = df[positive_flag_cols].sum(axis=1)
    df["Issue Summary"] = df.apply(_issue_summary, axis=1)

    for _, audit_row in df.iterrows():
        base_data = {
            "Record ID": audit_row["Record ID"],
            "Call Date": audit_row.get("Call Date"),
            "Audit Date": audit_row.get("Audit Date"),
            "Quality Auditor": audit_row.get("Quality Auditor"),
            "Call/Chat": audit_row.get("Call/Chat"),
            "Agent Name": audit_row.get("Agent Name"),
            "Supervisor name": audit_row.get("Supervisor name"),
            "LOB": audit_row.get("LOB"),
            "Sub-LOB": audit_row.get("Sub-LOB"),
            "Reason for call": audit_row.get("Reason for call"),
        }
        for parameter in PARAMETERS:
            long_rows.append(
                {
                    **base_data,
                    "Section": parameter.section,
                    "Parameter": parameter.name,
                    "Raw Response": audit_row.get(parameter.raw_col),
                    "Normalized Response": normalize_response(audit_row.get(parameter.raw_col)),
                    "Max Score": parameter.max_score,
                    "Score": audit_row[f"score__{parameter.name}"],
                    "Score %": audit_row[f"pct__{parameter.name}"],
                    "Status": audit_row[f"status__{parameter.name}"],
                    "Applicable": audit_row[f"applicable__{parameter.name}"],
                    "Failed": audit_row[f"failed__{parameter.name}"],
                    "Fatal": audit_row[f"fatal__{parameter.name}"],
                }
            )
        for flag_col in COMPLIANCE_FLAG_COLUMNS:
            compliance_rows.append(
                {
                    **base_data,
                    "Flag": flag_col,
                    "Raw Response": audit_row.get(flag_col),
                    "Normalized Response": normalize_response(audit_row.get(flag_col)),
                    "Positive": audit_row[f"Compliance Flag Positive :: {flag_col}"],
                }
            )

    long_df = pd.DataFrame(long_rows)
    compliance_df = pd.DataFrame(compliance_rows)

    percentage_columns = [
        "Overall Score (With Fatal)",
        "Overall Score (Without Fatal)",
        "Defect %",
        *[f"Section Score :: {section}" for section in RAW_SECTION_COLUMNS],
    ]
    for column in percentage_columns:
        if column in df.columns:
            df[column] = pd.to_numeric(df[column], errors="coerce")

    return df, long_df, compliance_df


def build_scored_dataset(workbook_source: str | Path | BytesIO | bytes) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = _excel_to_dataframe(workbook_source)
    return _build_scored_dataset_from_df(df)


def build_scored_dataset_from_rows(rows: list[list[Any]] | list[tuple[Any, ...]], start_row: int = 3) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = raw_rows_to_dataframe(rows, start_row=start_row)
    return _build_scored_dataset_from_df(df)


def summarize_dimension(df: pd.DataFrame, score_column: str, group_by: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    section_cols = [f"Section Score :: {section}" for section in RAW_SECTION_COLUMNS]
    summary = (
        df.groupby(group_by, dropna=False)
        .agg(
            Audits=("Record ID", "count"),
            Avg_Score=(score_column, "mean"),
            Avg_Defect=("Defect %", "mean"),
            Fatal_Rate=("Fatal Count", lambda s: float((s > 0).mean())),
            **{f"{section}": (col, "mean") for section, col in zip(RAW_SECTION_COLUMNS, section_cols)},
        )
        .reset_index()
    )
    return summary.sort_values(["Avg_Score", "Audits"], ascending=[False, False])


def top_findings(long_df: pd.DataFrame) -> pd.DataFrame:
    if long_df.empty:
        return pd.DataFrame(columns=["Parameter", "Status", "Observations"])
    findings = (
        long_df.loc[long_df["Status"].isin(["Fail", "Fatal", "Partial"])]
        .groupby(["Parameter", "Status"])
        .size()
        .reset_index(name="Observations")
        .sort_values(["Observations", "Parameter"], ascending=[False, True])
    )
    return findings


def parameter_failure_summary(long_df: pd.DataFrame) -> pd.DataFrame:
    if long_df.empty:
        return pd.DataFrame(columns=["Parameter", "Applicable Audits", "Failed Audits", "Failure Rate"])
    applicable = long_df[long_df["Applicable"]].groupby("Parameter").size().rename("Applicable Audits")
    failed = long_df[long_df["Status"].isin(["Fail", "Fatal", "Partial"])].groupby("Parameter").size().rename("Failed Audits")
    summary = pd.concat([applicable, failed], axis=1).fillna(0).reset_index()
    summary["Applicable Audits"] = summary["Applicable Audits"].astype(int)
    summary["Failed Audits"] = summary["Failed Audits"].astype(int)
    summary["Failure Rate"] = np.where(summary["Applicable Audits"] > 0, summary["Failed Audits"] / summary["Applicable Audits"], np.nan)
    return summary.sort_values(["Failure Rate", "Failed Audits"], ascending=[False, False])


def safe_percent(value: float | int | None) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return "-"
    return f"{value:.1%}"


def safe_number(value: float | int | None, decimals: int = 1) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return "-"
    return f"{value:,.{decimals}f}"
